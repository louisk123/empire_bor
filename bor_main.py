
import streamlit as st
import pdfplumber
import os
import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from rapidfuzz import process, fuzz



# Import all modules ONCE
from modules import (
    uae_vox,
    uae_galaxy,
    uae_safeer,
    uae_shaab,
    uae_star_cinemas,
    uae_cine_royale,
    uae_truth,
    uae_truth_weekly,
    kuwait_sky,
    kuwait_kncc,
    kuwait_ozone
)



module_map = {
    "UAE Vox":uae_vox,
    "UAE Star Cinemas": uae_star_cinemas,
    "UAE Cinepolis": uae_vox,
    "UAE Reel": uae_vox,
    "UAE NOVO": uae_vox,
    "UAE Cinemacity": uae_vox,
    "UAE Roxy": uae_vox,
    "UAE Cine Royale": uae_cine_royale,
    "UAE Galaxy": uae_galaxy,
    "UAE Truth": uae_truth,
    "UAE Truth Weekly": uae_truth_weekly,
    "UAE Shaab": uae_shaab,
    "UAE Safeer": uae_safeer,
    "Kuwait Vox": uae_vox,
    "Kuwait Hamra": uae_vox,
    "Kuwait Andalus": uae_vox,
    "Kuwait Sky":kuwait_sky,
    "Kuwait KNCC":kuwait_kncc
}

def add_cinema_movie_format_date(df,country_map):
    df["Country Code"] = df["Country"].map(country_map).fillna(df["Country"])
    df["Cinema_Movie_Format_Date"] = (df["Country Code"].astype(str) + " | " + df["Cinema"].astype(str) + " | " + df["Movie Mapped"].astype(str) + " | " + df["Format"].astype(str) + " | " +df["Date"].astype(str))
    return df



def find_last_real_row(ws):
    """
    Finds the last row that contains *any* data.
    It works even if there are empty rows in the middle.
    """
    last = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(v not in (None, "") for v in row):  # row has at least one value
            last = idx
    return last


def append_to_excel(excel_path, sheet_name, new_df):
    if len(new_df) == 0:
        #print("⚠️ new_df is EMPTY → nothing to append")
        return

    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    # find first free row
    last_row = find_last_real_row(ws)
    start_row = last_row + 1

    # write dataframe without overwriting anything
    for r_idx, row in new_df.iterrows():
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=start_row + r_idx, column=c_idx, value=value)

    wb.save(excel_path)
    #print("✔ Appended", len(new_df), "rows")


def get_sheet_name(row):

    week = row["Week Type"]
    summ = row["Is Summary"]




    if (week == "" or pd.isna(week)) and summ != 1:
        return "Daily BOR"
    if (week == "" or pd.isna(week)) and summ == 1:
        return "Daily BOR - summary"
    if week == "weekly" and summ != 1:
        return "Weekly BOR"
    if week == "weekly" and summ == 1:
        return "Weekly BOR - summary"
    return "Daily BOR"

def normalize_title(s: str) -> str:
    if not s or pd.isna(s):
        return ""

    s = str(s).lower()

    # remove language tags in [] and ()
    s = re.sub(r"\[.*?\]|\(.*?\)", " ", s)

    # replace 7 -> h ONLY when it behaves like a letter (start of token or inside token)
    s = re.sub(r"(?:(?<=\w)7(?=\w))|(?:\b7(?=[a-z]))", "h", s)

    # normalize fel/fi
    s = re.sub(r"\bfel\b", "fi", s)

    # remove punctuation
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()

    # token cleanup: remove leading "el" in arabic-translit tokens (elkhayal -> khayal)
    tokens = []
    for t in s.split():
        if t.startswith("el") and len(t) > 4:
            t = t[2:]
        tokens.append(t)
    return " ".join(tokens)


def map_movie(name, movie_list, threshold=80):
    if not name or pd.isna(name):
        return name

    # build normalized lookup ONCE per call
    norm_to_original = {normalize_title(m): m for m in movie_list}

    name_norm = normalize_title(name)

    match = process.extractOne(
        name_norm,
        norm_to_original.keys(),
        scorer=fuzz.token_set_ratio
    )

    if match and match[1] >= threshold:
        return norm_to_original[match[0]]

    return name

def screen_rule(g):
    valid_screens = (
        g.dropna(subset=["Screen", "Time"])
         .groupby("Screen")["Time"]
         .nunique()
    )
    n = (valid_screens >= 3).sum()
    return n if n > 0 else 1


def fix_dates(file_df):

    flip_exhibitors = {"Cinepolis", "Vox", "Reel", "NOVO", "Cinemacity", "Roxy"}
    slash_exhibitors = {"Galaxy", "Truth", "Truth Weekly", "Shaab", "Safeer"}

    # ensure string
    file_df["Date"] = file_df["Date"].astype(str)

    # 1) replace - with / for slash exhibitors
    mask_slash = file_df["Exhibitor"].isin(slash_exhibitors)
    file_df.loc[mask_slash, "Date"] = file_df.loc[mask_slash, "Date"].str.replace("-", "/", regex=False)

    # 2) flip day/month for flip exhibitors when not summary
    mask_flip = (
        file_df["Exhibitor"].isin(flip_exhibitors) &
        (file_df["Is Summary"] != 1)
    )

    # split and reassemble safely
    date_parts = file_df.loc[mask_flip, "Date"].str.split("/", expand=True)
    file_df.loc[mask_flip, "Date"] = (
        date_parts[1] + "/" + date_parts[0] + "/" + date_parts[2]
    )
    return file_df



def fix_dates1(file_df, date_format_map):

    FORMAT_MAP = {
        "yyyy-mm-dd": ["%Y-%m-%d", "%y-%m-%d"],
        "dd/mm/yyyy": ["%d/%m/%Y", "%d/%m/%y"],
        "mm/dd/yyyy": ["%m/%d/%Y", "%m/%d/%y"],
        "dd-mm-yyyy": ["%d-%m-%Y", "%d-%m-%y"],
        "dd/mon/yyyy": ["%d/%b/%Y", "%d/%b/%y", "%d/%B/%Y", "%d/%B/%y"],
        "dd-mon-yyyy": ["%d-%b-%Y", "%d-%b-%y", "%d-%B-%Y", "%d-%B-%y"]
    }

    def convert(date_str, current_format):
        if pd.isna(date_str):
            return date_str

        if current_format not in FORMAT_MAP:
            return date_str

        s = str(date_str).strip()

        for fmt in FORMAT_MAP[current_format]:
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime("%d/%m/%Y")
            except ValueError:
                pass

        return s

    file_df["_KEY"] = (
        file_df["Cinema"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    mask = file_df["Is Summary"].fillna(0).astype(int) != 1

    file_df.loc[mask, "Date"] = file_df.loc[mask].apply(
        lambda r: convert(
            r["Date"],
            date_format_map.get(r["_KEY"])
        ),
        axis=1
    )

    file_df.drop(columns="_KEY", inplace=True)
    return file_df



def get_first_line(file_path,cinema_map):  #chekc if file is pdf or excel and treat differently
    ext = os.path.splitext(file_path)[1].lower()
    text=""
    # PDF
    if ext == ".pdf":
        try:
            with pdfplumber.open(file_path) as pdf:
                text = pdf.pages[0].extract_text() or ""
                return text.split("\n")[0].strip()
        except:
            return None

    # EXCEL
    elif ext in [".xlsx", ".xls"]:
        try:
            df = pd.read_excel(file_path, header=None)
            cell_value = str(df.iloc[6, 1]).strip()
            if cell_value.upper() in cinema_map:
                return cell_value   # matched → return mapped value
            else:
                return None                     # not found
        except:
            return None

    return None




def process_pdf(pdf_path, excel_path):


    #file_df=pd.DataFrame()
    now_value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # read the mapping sheet
    mapping_df = pd.read_excel(
        excel_path,
        sheet_name="Cinemas Mapping",
        usecols=["Name from File", "Line", "Exhibitor", "Country","BOR File",	"BOR Exhibitor","File Date Format"]
    )
    movies_df = pd.read_excel(
        excel_path,
        sheet_name="Movies",
        usecols=["BOR Movie Name"]
    )
    
    
    movie_list = (
        movies_df["BOR Movie Name"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )
    
    format_df = pd.read_excel(
    excel_path,
    sheet_name="Format",
    usecols=["PDF", "BOR"]
    )
    
    format_map = (
    format_df
    .dropna(subset=["PDF", "BOR"])
    .assign(
        PDF=lambda d: d["PDF"].astype(str).str.strip().str.upper()
    )
    .set_index("PDF")["BOR"]
    .to_dict()
    )
    


    cinema_df = pd.read_excel(
        excel_path,
        sheet_name="Cinemas Mapping",
        usecols=["Name from File", "BOR File", "BOR Exhibitor"]
    )
    
    cinema_map = (
        cinema_df
        .dropna(subset=["Name from File", "BOR File"])
        .assign(
            PDF=lambda d: d["Name from File"].astype(str).str.strip().str.upper()
        )
        .set_index("PDF")["BOR File"]
        .to_dict()
    )

    
    exhibitor_map = (
        cinema_df
        .dropna(subset=["Name from File", "BOR Exhibitor"])
        .assign(
            PDF=lambda d: d["Name from File"].astype(str).str.strip().str.upper()
        )
        .set_index("PDF")["BOR Exhibitor"]
        .to_dict()
    )

    date_format_map = (
        mapping_df
        .dropna(subset=["Name from File", "File Date Format"])
        .assign(
            KEY=lambda d: d["Name from File"].astype(str).str.strip().str.upper()
        )
        .set_index("KEY")["File Date Format"]
        .to_dict()
    )
    
    country_map = {
    "Lebanon": "LB",
    "Jordan": "JO",
    "Iraq": "IQ",
    "Syria": "SY",
    "Egypt": "EG",
    "Ethiopia": "ET",
    "UAE": "AE",
    "Kuwait": "KW",
    "Qatar": "QR",
    "Bahrain": "BH",
    "Oman": "OM",
    "KSA": "SA",
    "Djibouti": "DJ",
    "Palestine": "PA",
}

    


    first_line = get_first_line(pdf_path,cinema_map)

    if first_line is None:
        return


    # Find matching cinema
    cinema_found = None
    cinema_country = None
    exhibitor = None
    first_line_upper = first_line.upper()

    for _, row in mapping_df.iterrows():

        if first_line_upper == "AL MARIAH MALL ABU DHABHI":
            first_line_upper = first_line_upper + " " + text.split("\n")[1].strip().upper()
            print(first_line_upper)

        cinema_name = str(row["Name from File"]).upper()
        if cinema_name in first_line_upper:
            cinema_found = cinema_name
            cinema_country = row["Country"]
            exhibitor = row["Exhibitor"]
            st.write("cinemae foud")
            break

    if cinema_found is None:
        print("No cinema match for:", pdf_path)
        return

    # Call correct module
    key = f"{cinema_country} {exhibitor}".strip()
    module = module_map.get(key)


    try:
        if exhibitor=="KNCC":
            file_df = module.fetch_data(pdf_path, exhibitor,cinema_map)
        else:
            file_df = module.fetch_data(pdf_path, exhibitor)

        if file_df is None or len(file_df) == 0:
            print("Empty df, skipping:", pdf_path)
            return

        # Fix column names
        file_df.columns = [
                    "File", "Exhibitor","Cinema", "Week Type", "Extraction Date",
                    "Movie","Date","Time","Screen","Format","Ticket Type",
                    "Admits","Gross","Net","Comp","Is Summary","Summary Sessions"
          ]

        file_df["Extraction Date"] = now_value
        file_df["Country"] = cinema_country
        
        file_df["Movie Mapped"] = file_df["Movie"].apply(
            lambda x: map_movie(x, movie_list)
        )

        #file_df=fix_dates(file_df)
        file_df = fix_dates1(file_df, date_format_map)

        file_df["Format"] = (
            file_df["Format"]
            .astype(str)
            .str.strip()
            .str.upper()
            .map(format_map)
            .fillna(file_df["Format"])  # keep original if not found
        )

        file_df["Cinema"] = file_df["Cinema"].astype(str).str.strip().str.upper()
        file_df["Exhibitor"] = file_df["Exhibitor"].astype(str).str.strip().str.upper()

        # Replace Exhibitor → BOR Exhibitor
        file_df["Exhibitor"] = (
            file_df["Cinema"]
            .map(exhibitor_map)
            .fillna(file_df["Exhibitor"])
        )

        
        # Replace Cinema → BOR File
        file_df["Cinema"] = (
            file_df["Cinema"]
            .map(cinema_map)
            .fillna(file_df["Cinema"])
        )
        



        EXPECTED_ORDER = [
              "File","Exhibitor","Cinema","Week Type","Extraction Date",
              "Movie","Movie Mapped","Date","Time","Screen","Format",
              "Ticket Type","Admits","Gross","Net","Comp",
              "Is Summary","Summary Sessions","Country"
              ]

        file_df = file_df.reindex(columns=EXPECTED_ORDER)

        append_to_excel(excel_path, "Raw Data", file_df)

        # Normalize fields
        file_df["Week Type"] = file_df["Week Type"].fillna("").str.strip().str.lower()
        file_df["Is Summary"] = file_df["Is Summary"].fillna(0).astype(int)
        
        #compute number of screens
        group_cols = [
                  "File","Exhibitor","Cinema","Week Type","Extraction Date",
              "Movie","Movie Mapped","Date",
              "Is Summary","Country","Format"
        ]

        screen_counts = file_df.groupby(group_cols).apply(screen_rule)

        file_df["Screen_Calc"] = (
                  file_df
                  .set_index(group_cols)
                  .index
                  .map(screen_counts)
        )

        
        
        
        
        # --------------------------
        # SPLIT DATA
        # --------------------------

        daily_df = file_df[
            (file_df["Week Type"] != "weekly") &
            (file_df["Is Summary"] != 1)
        ]

        daily_sum_df = file_df[
            (file_df["Week Type"] != "weekly") &
            (file_df["Is Summary"] == 1)
        ]

        weekly_df = file_df[
            (file_df["Week Type"] == "weekly") &
            (file_df["Is Summary"] != 1)
        ]

        weekly_sum_df = file_df[
            (file_df["Week Type"] == "weekly") &
            (file_df["Is Summary"] == 1)
        ]

        # --------------------------
        # AGGREGATIONS
        # --------------------------

        group_cols = [
            "Country", "File", "Exhibitor", "Cinema", "Extraction Date",
            "Movie","Movie Mapped", "Date", "Format"
        ]

        agg_rules = {
            "Admits": "sum",
            "Gross": "sum",
            "Net": "sum",
            "Comp": "sum",
            "Summary Sessions": "sum",
            "Screen_Calc": "max",
            "Time": "nunique"
        }

        daily_agg = daily_df.groupby(group_cols).agg(agg_rules).reset_index()
        daily_sum_agg = daily_sum_df.groupby(group_cols).agg(agg_rules).reset_index()
        weekly_agg = weekly_df.groupby(group_cols).agg(agg_rules).reset_index()
        weekly_sum_agg = weekly_sum_df.groupby(group_cols).agg(agg_rules).reset_index()
        
        daily_agg = add_cinema_movie_format_date(daily_agg,country_map)
        daily_sum_agg = add_cinema_movie_format_date(daily_sum_agg,country_map)
        weekly_agg = add_cinema_movie_format_date(weekly_agg,country_map)
        weekly_sum_agg = add_cinema_movie_format_date(weekly_sum_agg,country_map)



        # Write results
        append_to_excel(excel_path, "Daily BOR", daily_agg)
        append_to_excel(excel_path, "Daily BOR - Summary", daily_sum_agg)
        append_to_excel(excel_path, "Weekly BOR", weekly_agg)
        append_to_excel(excel_path, "Weekly BOR - Summary", weekly_sum_agg)

    except Exception as e:
        print("Error calling module:", e)
