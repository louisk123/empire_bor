import streamlit as st
import pandas as pd
import tempfile
import os
from openpyxl import load_workbook
from bor_main import process_pdf  
from bor_main import process_excel 

st.title("Empire BOR Extraction System")
st.write("Upload one Excel file, multiple PDF files, and optional BOR Excel files.")

# --------------------------
# UPLOAD EXCEL OUTPUT FILE
# --------------------------
excel_file = st.file_uploader(
    "Upload Excel Output File",
    type=["xlsx"]
)

# --------------------------
# UPLOAD PDF FILES
# --------------------------
pdf_files = st.file_uploader(
    "Upload  Files",
    type=["pdf","xlsx"],
    accept_multiple_files=True
)

# --------------------------
# UPLOAD BOR EXCEL FILES (OPTIONAL)
# --------------------------
bor_excel_files = st.file_uploader(
    "Upload BOR Files (Excel)",
    type=["xlsm", "xlsx"],
    accept_multiple_files=True
)

# --------------------------
# RUN PROCESSING
# --------------------------
if st.button("Start Processing"):

    if not excel_file:
        st.error("Please upload an Excel file.")
        st.stop()

    if not pdf_files:
        st.error("Please upload at least one file.")
        st.stop()

    # Save Excel temporarily
    temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp_excel.write(excel_file.read())
    temp_excel.close()
    temp_excel_path = temp_excel.name

    st.success("Excel loaded successfully.")

    # --------------------------
    # PROCESS PDF FILES
    # --------------------------
    progress = st.progress(0)
    total = len(pdf_files)

    for idx, pdf in enumerate(pdf_files, start=1):
        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        temp_pdf.write(pdf.read())
        temp_pdf.close()
        temp_pdf_path = temp_pdf.name

        if suffix == ".pdf":
            process_pdf(temp_path, temp_excel_path)
        else:
            process_excel(temp_path, temp_excel_path)

        os.remove(temp_pdf_path)
        progress.progress(idx / total)

    # --------------------------
    # COPY DATA FROM BOR FILES (ONCE)
    # --------------------------
    if bor_excel_files:
        wb_target = load_workbook(temp_excel_path)
        ws_target = wb_target["Data From BOR"]
    
        start_row = ws_target.max_row + 1
    
        for bor_file in bor_excel_files:
            tmp_bor = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp_bor.write(bor_file.read())
            tmp_bor.close()
            tmp_bor_path = tmp_bor.name
    
            # READ USING PANDAS (KEEP ORIGINAL TYPES)
            df_bor = pd.read_excel(
                tmp_bor_path,
                sheet_name="Data"
            )
    
            # Convert ONLY Screening Date
            if "Screening Date" in df_bor.columns:
                df_bor["Screening Date"] = (
                    pd.to_datetime(df_bor["Screening Date"], errors="coerce")
                      .dt.strftime("%d/%m/%Y")
                )

            # Skip header row, keep original 16 + new column
            df_bor = df_bor.iloc[1:, :17]

                    # Add combined column (BOR context only)
            df_bor["Cinema_Movie_Format_Date"] = (
            df_bor["Territory"].astype(str) + " | " +
            df_bor["Theater"].astype(str) + " | " +
            df_bor["Description"].astype(str) + " | " +
            df_bor["Movie Type"].astype(str) + " | " +
            df_bor["Screening Date"].astype(str)
            )


            
    
            # Append rows (fast)
            for row in df_bor.itertuples(index=False, name=None):
                ws_target.append(row)
    
            os.remove(tmp_bor_path)
    
        wb_target.save(temp_excel_path)



    st.success("Processing complete!")

    # --------------------------
    # DOWNLOAD RESULT
    # --------------------------
    with open(temp_excel_path, "rb") as f:
        st.download_button(
            label="Download Updated Excel",
            data=f,
            file_name="Updated_BOR_Output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    os.remove(temp_excel_path)
